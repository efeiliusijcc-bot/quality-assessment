"""
电子元器件装配质量评估系统 - 测试数据生成脚本
运行: python generate_test_data.py
输出: test_data/ 目录下的 Excel 文件
依赖: pip install openpyxl
"""

import random
import uuid
import os
from datetime import datetime, timedelta
from openpyxl import Workbook

random.seed(42)
OUTPUT_DIR = "test_data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# 基础数据定义
# ============================================================

PROCESS_STEPS = [
    {"stepCode": "SMT_MOUNT", "stepName": "SMT贴装", "stepOrder": 1, "isInspection": False},
    {"stepCode": "REFLOW", "stepName": "回流焊", "stepOrder": 2, "isInspection": False},
    {"stepCode": "AOI", "stepName": "AOI光学检测", "stepOrder": 3, "isInspection": True},
    {"stepCode": "XRAY", "stepName": "X-Ray检测", "stepOrder": 4, "isInspection": True},
    {"stepCode": "ICT", "stepName": "ICT在线测试", "stepOrder": 5, "isInspection": True},
    {"stepCode": "FCT", "stepName": "FCT功能测试", "stepOrder": 6, "isInspection": True},
    {"stepCode": "CONFORMAL", "stepName": "三防涂覆", "stepOrder": 7, "isInspection": False},
    {"stepCode": "FINAL_INSP", "stepName": "终检", "stepOrder": 8, "isInspection": True},
]

WORKSTATIONS = [
    {"stationCode": "WS-01", "stationName": "贴装工位A1", "location": "1号线-东侧"},
    {"stationCode": "WS-02", "stationName": "贴装工位A2", "location": "1号线-西侧"},
    {"stationCode": "WS-03", "stationName": "焊接工位B1", "location": "2号线-东侧"},
    {"stationCode": "WS-04", "stationName": "焊接工位B2", "location": "2号线-西侧"},
    {"stationCode": "WS-05", "stationName": "检测工位C1", "location": "3号线-东侧"},
    {"stationCode": "WS-06", "stationName": "检测工位C2", "location": "3号线-西侧"},
    {"stationCode": "WS-07", "stationName": "装配工位D1", "location": "4号线-东侧"},
    {"stationCode": "WS-08", "stationName": "装配工位D2", "location": "4号线-西侧"},
    {"stationCode": "WS-09", "stationName": "涂覆工位E1", "location": "5号线"},
    {"stationCode": "WS-10", "stationName": "终检工位F1", "location": "6号线"},
]

EQUIPMENT_LIST = [
    {"equipmentCode": "EQ-SMT-001", "equipmentName": "高速贴片机-1", "equipmentType": "贴片机", "manufacturer": "Yamaha", "modelNo": "YSM20R"},
    {"equipmentCode": "EQ-SMT-002", "equipmentName": "高速贴片机-2", "equipmentType": "贴片机", "manufacturer": "Fuji", "modelNo": "NXT III"},
    {"equipmentCode": "EQ-REFLOW-001", "equipmentName": "回流焊炉-1", "equipmentType": "回流焊", "manufacturer": "Heller", "modelNo": "1913MK5"},
    {"equipmentCode": "EQ-REFLOW-002", "equipmentName": "回流焊炉-2", "equipmentType": "回流焊", "manufacturer": "BTU", "modelNo": "Pyramax"},
    {"equipmentCode": "EQ-AOI-001", "equipmentName": "AOI检测仪-1", "equipmentType": "AOI", "manufacturer": "KohYoung", "modelNo": "Zenith"},
    {"equipmentCode": "EQ-XRAY-001", "equipmentName": "X-Ray检测仪-1", "equipmentType": "X-Ray", "manufacturer": "Nikon", "modelNo": "XT V 160"},
    {"equipmentCode": "EQ-ICT-001", "equipmentName": "ICT测试仪-1", "equipmentType": "ICT", "manufacturer": "Keysight", "modelNo": "i3070"},
    {"equipmentCode": "EQ-FCT-001", "equipmentName": "FCT测试台-1", "equipmentType": "FCT", "manufacturer": "NI", "modelNo": "PXIe-1082"},
]

PRODUCT_TYPES = [
    {"productCode": "PCB-A001", "productName": "主控板-V3", "materialSystem": "无铅SAC305", "specification": "6层板 160x120mm"},
    {"productCode": "PCB-B002", "productName": "电源模块-V2", "materialSystem": "无铅SAC305", "specification": "4层板 100x80mm"},
    {"productCode": "PCB-C003", "productName": "信号处理板-V1", "materialSystem": "有铅Sn63Pb37", "specification": "8层板 200x150mm"},
    {"productCode": "PCB-D004", "productName": "通信接口板-V2", "materialSystem": "无铅SAC305", "specification": "6层板 120x90mm"},
    {"productCode": "PCB-E005", "productName": "传感器板-V1", "materialSystem": "无铅SAC305", "specification": "4层板 60x40mm"},
]

PARAM_DEFS = [
    {"paramCode": "PREHEAT_TEMP", "paramName": "预热温度", "paramCategory": "温度", "dataType": "NUMBER", "unit": "℃", "lowerLimit": 140, "upperLimit": 170, "standardValue": 155},
    {"paramCode": "REFLOW_TEMP", "paramName": "回流峰值温度", "paramCategory": "温度", "dataType": "NUMBER", "unit": "℃", "lowerLimit": 220, "upperLimit": 250, "standardValue": 235},
    {"paramCode": "BELT_SPEED", "paramName": "传送带速度", "paramCategory": "速度", "dataType": "NUMBER", "unit": "cm/min", "lowerLimit": 80, "upperLimit": 100, "standardValue": 90},
    {"paramCode": "O2_PPM", "paramName": "氧含量", "paramCategory": "气氛", "dataType": "NUMBER", "unit": "ppm", "lowerLimit": 400, "upperLimit": 600, "standardValue": 500},
    {"paramCode": "HUMIDITY", "paramName": "环境湿度", "paramCategory": "环境", "dataType": "NUMBER", "unit": "%RH", "lowerLimit": 30, "upperLimit": 60, "standardValue": 45},
    {"paramCode": "CURRENT", "paramName": "焊接电流", "paramCategory": "电气", "dataType": "NUMBER", "unit": "A", "lowerLimit": 1.0, "upperLimit": 1.5, "standardValue": 1.2},
    {"paramCode": "PRESSURE", "paramName": "贴装压力", "paramCategory": "力学", "dataType": "NUMBER", "unit": "N", "lowerLimit": 3.5, "upperLimit": 5.0, "standardValue": 4.2},
    {"paramCode": "PLACEMENT_OFFSET", "paramName": "贴装偏移量", "paramCategory": "精度", "dataType": "NUMBER", "unit": "mm", "lowerLimit": 0, "upperLimit": 0.15, "standardValue": 0.05},
]

DEFECT_TYPES = [
    {"defectCode": "SOLDER_VOID", "defectName": "虚焊", "defectCategory": "焊接缺陷", "defaultSeverity": "严重"},
    {"defectCode": "BRIDGE", "defectName": "桥连", "defectCategory": "焊接缺陷", "defaultSeverity": "严重"},
    {"defectCode": "TOMBSTONE", "defectName": "立碑", "defectCategory": "贴装缺陷", "defaultSeverity": "中等"},
    {"defectCode": "OFFSET", "defectName": "偏移", "defectCategory": "贴装缺陷", "defaultSeverity": "中等"},
    {"defectCode": "MISSING", "defectName": "缺件", "defectCategory": "贴装缺陷", "defaultSeverity": "严重"},
    {"defectCode": "CRACK", "defectName": "裂纹", "defectCategory": "焊接缺陷", "defaultSeverity": "严重"},
    {"defectCode": "CONTAMINATION", "defectName": "污染", "defectCategory": "外观缺陷", "defaultSeverity": "轻微"},
    {"defectCode": "SCRATCH", "defectName": "划伤", "defectCategory": "外观缺陷", "defaultSeverity": "轻微"},
]

QUALITY_METRICS = [
    {"metricCode": "SOLDER_COVERAGE", "metricName": "焊点覆盖率", "unit": "%", "lowerLimit": 95, "upperLimit": 100, "targetValue": 98, "passRule": ">=95", "severityWeight": 0.3},
    {"metricCode": "PLACEMENT_ACCURACY", "metricName": "贴装精度", "unit": "mm", "lowerLimit": 0, "upperLimit": 0.1, "targetValue": 0.05, "passRule": "<=0.1", "severityWeight": 0.25},
    {"metricCode": "SHEAR_STRENGTH", "metricName": "剪切强度", "unit": "N", "lowerLimit": 20, "upperLimit": 50, "targetValue": 35, "passRule": ">=20", "severityWeight": 0.25},
    {"metricCode": "INSULATION_RESISTANCE", "metricName": "绝缘电阻", "unit": "MΩ", "lowerLimit": 100, "upperLimit": 10000, "targetValue": 1000, "passRule": ">=100", "severityWeight": 0.2},
]

BATCH_NO_PREFIXES = ["BATCH-2401", "BATCH-2402", "BATCH-2403", "BATCH-2404", "BATCH-2405",
                     "BATCH-2406", "BATCH-2407", "BATCH-2408", "BATCH-2409", "BATCH-2410"]

STATIONS_CN = ["贴装工位A1", "贴装工位A2", "焊接工位B1", "焊接工位B2", "检测工位C1", "检测工位C2", "装配工位D1", "装配工位D2"]


def uid():
    return str(uuid.uuid4())


def rand_time(base_day=0):
    """生成随机时间，基于 2026-04 月份"""
    d = datetime(2026, 4, 1) + timedelta(days=random.randint(0, 28) + base_day)
    h = random.randint(7, 19)
    m = random.randint(0, 59)
    s = random.randint(0, 59)
    return d.replace(hour=h, minute=m, second=s)


def rand_value(lower, upper, decimals=2):
    """在范围内生成随机值"""
    return round(random.uniform(lower, upper), decimals)


def save_workbook(wb, filename):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  [OK] {path}")


# ============================================================
# 生成 core 模块数据
# ============================================================

def gen_core_data():
    print("生成 core 模块数据...")
    wb = Workbook()

    # ProcessStep
    ws = wb.active
    ws.title = "process_step"
    ws.append(["stepId", "stepCode", "stepName", "stepOrder", "isInspection", "description", "createdAt"])
    step_ids = []
    for s in PROCESS_STEPS:
        sid = uid()
        step_ids.append(sid)
        ws.append([sid, s["stepCode"], s["stepName"], s["stepOrder"], s["isInspection"],
                   f"{s['stepName']}工序", rand_time().isoformat()])

    # Workstation
    ws2 = wb.create_sheet("workstation")
    ws2.append(["stationId", "stepId", "stationCode", "stationName", "location", "status", "createdAt"])
    station_ids = []
    for i, w in enumerate(WORKSTATIONS):
        wid = uid()
        station_ids.append(wid)
        step_id = step_ids[i % len(step_ids)]
        ws2.append([wid, step_id, w["stationCode"], w["stationName"], w["location"], "ACTIVE", rand_time().isoformat()])

    # Equipment
    ws3 = wb.create_sheet("equipment")
    ws3.append(["equipmentId", "stationId", "equipmentCode", "equipmentName", "equipmentType",
                 "manufacturer", "modelNo", "status", "installedAt", "createdAt"])
    equip_ids = []
    for i, e in enumerate(EQUIPMENT_LIST):
        eid = uid()
        equip_ids.append(eid)
        sid = station_ids[i % len(station_ids)]
        ws3.append([eid, sid, e["equipmentCode"], e["equipmentName"], e["equipmentType"],
                     e["manufacturer"], e["modelNo"], "ACTIVE",
                     (datetime(2025, 1, 1) + timedelta(days=random.randint(0, 365))).strftime("%Y-%m-%d"),
                     rand_time().isoformat()])

    # ProductType
    ws4 = wb.create_sheet("product_type")
    ws4.append(["productTypeId", "productCode", "productName", "materialSystem", "specification", "createdAt"])
    product_ids = []
    for p in PRODUCT_TYPES:
        pid = uid()
        product_ids.append(pid)
        ws4.append([pid, p["productCode"], p["productName"], p["materialSystem"], p["specification"], rand_time().isoformat()])

    # ParameterDef
    ws5 = wb.create_sheet("parameter_def")
    ws5.append(["paramId", "stepId", "paramCode", "paramName", "paramCategory", "dataType", "unit",
                 "lowerLimit", "upperLimit", "standardValue", "requiredFlag", "description", "createdAt"])
    param_ids = []
    for p in PARAM_DEFS:
        pid = uid()
        param_ids.append(pid)
        step_id = step_ids[1] if p["paramCode"] in ["PREHEAT_TEMP", "REFLOW_TEMP", "BELT_SPEED", "O2_PPM"] else step_ids[0]
        ws5.append([pid, step_id, p["paramCode"], p["paramName"], p["paramCategory"], p["dataType"], p["unit"],
                     p["lowerLimit"], p["upperLimit"], p["standardValue"], True,
                     f"{p['paramName']}参数", rand_time().isoformat()])

    save_workbook(wb, "core_data.xlsx")
    return step_ids, station_ids, equip_ids, product_ids, param_ids


# ============================================================
# 生成 prod 模块数据
# ============================================================

def gen_prod_data(step_ids, station_ids, equip_ids, product_ids, param_ids):
    print("生成 prod 模块数据...")
    wb = Workbook()

    # ProductionBatch
    ws = wb.active
    ws.title = "production_batch"
    ws.append(["batchId", "batchNo", "productTypeId", "planQty", "actualQty", "startTime", "endTime", "batchStatus", "createdBy", "createdAt"])
    batch_ids = []
    for i, bno in enumerate(BATCH_NO_PREFIXES):
        bid = uid()
        batch_ids.append(bid)
        pid = product_ids[i % len(product_ids)]
        plan = random.choice([50, 100, 200, 500])
        actual = plan - random.randint(0, 10)
        start = rand_time(i * 3)
        end = start + timedelta(hours=random.randint(4, 12))
        status = random.choice(["COMPLETED", "COMPLETED", "COMPLETED", "IN_PROGRESS"])
        ws.append([bid, bno, pid, plan, actual if status == "COMPLETED" else None,
                   start.isoformat(), end.isoformat() if status == "COMPLETED" else None,
                   status, uid(), rand_time().isoformat()])

    # ProductUnit
    ws2 = wb.create_sheet("product_unit")
    ws2.append(["unitId", "batchId", "serialNo", "currentStepId", "unitStatus", "createdAt"])
    unit_ids = []
    for i, bid in enumerate(batch_ids):
        for j in range(random.randint(8, 15)):
            unid = uid()
            unit_ids.append((unid, bid))
            serial = f"U{i+1:02d}-{j+1:03d}"
            ws2.append([unid, bid, serial, random.choice(step_ids), random.choice(["PASS", "PASS", "PASS", "FAIL"]), rand_time(i * 3).isoformat()])

    # ProcessRun
    ws3 = wb.create_sheet("process_run")
    ws3.append(["runId", "batchId", "unitId", "stepId", "stationId", "equipmentId", "recipeId",
                 "operatorId", "runNo", "startTime", "endTime", "runStatus", "createdAt"])
    run_ids = []
    for i, (unid, bid) in enumerate(unit_ids[:80]):  # 取前80个单元
        for step_id in random.sample(step_ids, min(3, len(step_ids))):
            rid = uid()
            run_ids.append((rid, unid, bid, step_id))
            start = rand_time(i)
            end = start + timedelta(minutes=random.randint(5, 30))
            ws3.append([rid, bid, unid, step_id, random.choice(station_ids), random.choice(equip_ids),
                        None, uid(), f"R-{i:04d}", start.isoformat(), end.isoformat(),
                        "COMPLETED", rand_time().isoformat()])

    # ParameterValue
    ws4 = wb.create_sheet("parameter_value")
    ws4.append(["valueId", "runId", "paramId", "measuredAt", "valueNum", "qualityFlag", "createdAt"])
    for rid, _, _, _ in run_ids[:50]:
        for pid in random.sample(param_ids, min(4, len(param_ids))):
            pdef = PARAM_DEFS[param_ids.index(pid) % len(PARAM_DEFS)]
            val = rand_value(pdef["lowerLimit"], pdef["upperLimit"])
            flag = "NORMAL" if pdef["lowerLimit"] <= val <= pdef["upperLimit"] else "WARNING"
            ws4.append([uid(), rid, pid, rand_time().isoformat(), val, flag, rand_time().isoformat()])

    save_workbook(wb, "prod_data.xlsx")
    return batch_ids, unit_ids, run_ids


# ============================================================
# 生成 qc 模块数据
# ============================================================

def gen_qc_data(step_ids, unit_ids, run_ids):
    print("生成 qc 模块数据...")
    wb = Workbook()

    # DefectType
    ws = wb.active
    ws.title = "defect_type"
    ws.append(["defectTypeId", "stepId", "defectCode", "defectName", "defectCategory", "defaultSeverity", "description", "createdAt"])
    defect_type_ids = []
    for d in DEFECT_TYPES:
        did = uid()
        defect_type_ids.append(did)
        ws.append([did, random.choice(step_ids), d["defectCode"], d["defectName"],
                   d["defectCategory"], d["defaultSeverity"], f"{d['defectName']}缺陷", rand_time().isoformat()])

    # QualityMetricDef
    ws2 = wb.create_sheet("quality_metric_def")
    ws2.append(["metricId", "stepId", "metricCode", "metricName", "unit", "lowerLimit", "upperLimit",
                 "targetValue", "passRule", "severityWeight", "description", "createdAt"])
    metric_ids = []
    for m in QUALITY_METRICS:
        mid = uid()
        metric_ids.append(mid)
        ws2.append([mid, random.choice(step_ids), m["metricCode"], m["metricName"], m["unit"],
                     m["lowerLimit"], m["upperLimit"], m["targetValue"], m["passRule"],
                     m["severityWeight"], f"{m['metricName']}指标", rand_time().isoformat()])

    # InspectionTask
    ws3 = wb.create_sheet("inspection_task")
    ws3.append(["inspectionId", "runId", "unitId", "stepId", "inspectionType", "modelName", "modelVersion",
                 "resultStatus", "confidence", "inspectedAt", "createdAt"])
    inspection_ids = []
    for i, (rid, unid, bid, step_id) in enumerate(run_ids[:60]):
        iid = uid()
        inspection_ids.append(iid)
        is_pass = random.random() > 0.15
        conf = rand_value(0.85, 0.99, 4) if is_pass else rand_value(0.5, 0.85, 4)
        ws3.append([iid, rid, unid, step_id, random.choice(["AOI", "XRAY", "ICT", "FCT"]),
                    "ResNet-50", "v2.1", "PASS" if is_pass else "FAIL", conf,
                    rand_time().isoformat(), rand_time().isoformat()])

    # DefectRecord
    ws4 = wb.create_sheet("defect_record")
    ws4.append(["defectId", "inspectionId", "unitId", "defectTypeId", "defectCount", "confidence",
                 "severityLevel", "isCritical", "createdAt"])
    for iid in inspection_ids:
        if random.random() < 0.3:  # 30%的检测任务有缺陷
            did = uid()
            dtid = random.choice(defect_type_ids)
            sev = random.choice(["轻微", "中等", "严重"])
            ws4.append([did, iid, uid(), dtid, random.randint(1, 5),
                        rand_value(0.7, 0.98, 4), sev, sev == "严重", rand_time().isoformat()])

    # QualityMeasurement
    ws5 = wb.create_sheet("quality_measurement")
    ws5.append(["measurementId", "runId", "unitId", "metricId", "measuredAt", "valueNum",
                 "isPass", "deviationValue", "measurementMethod", "createdAt"])
    for rid, unid, bid, step_id in run_ids[:40]:
        for mid in random.sample(metric_ids, min(2, len(metric_ids))):
            mdef = QUALITY_METRICS[metric_ids.index(mid) % len(QUALITY_METRICS)]
            val = rand_value(mdef["lowerLimit"] * 0.9, mdef["upperLimit"] * 1.1)
            is_pass = mdef["lowerLimit"] <= val <= mdef["upperLimit"]
            dev = abs(val - mdef["targetValue"])
            ws5.append([uid(), rid, unid, mid, rand_time().isoformat(), val,
                        is_pass, round(dev, 4), "自动检测", rand_time().isoformat()])

    save_workbook(wb, "qc_data.xlsx")
    return defect_type_ids, metric_ids


# ============================================================
# 生成 kg 模块数据
# ============================================================

def gen_kg_data(step_ids, param_ids, defect_type_ids, batch_ids):
    print("生成 kg 模块数据...")
    wb = Workbook()

    # GraphVersion
    ws = wb.active
    ws.title = "graph_version"
    ws.append(["graphVersionId", "graphName", "versionNo", "description", "createdAt"])
    gvid = uid()
    ws.append([gvid, "电子元器件装配知识图谱", "v1.0", "初始版本知识图谱", rand_time().isoformat()])

    # KgEntity
    ws2 = wb.create_sheet("kg_entity")
    ws2.append(["entityId", "graphVersionId", "entityType", "refSchema", "refTable", "refId",
                 "entityCode", "entityName", "createdAt"])
    entity_map = {}  # entityType -> [entityId]

    # 工序实体
    for i, sid in enumerate(step_ids):
        eid = uid()
        entity_map.setdefault("ProcessStep", []).append(eid)
        ws2.append([eid, gvid, "ProcessStep", "core", "process_step", sid,
                     PROCESS_STEPS[i]["stepCode"], PROCESS_STEPS[i]["stepName"], rand_time().isoformat()])

    # 参数实体
    for i, pid in enumerate(param_ids):
        eid = uid()
        entity_map.setdefault("ProcessParameter", []).append(eid)
        ws2.append([eid, gvid, "ProcessParameter", "core", "parameter_def", pid,
                     PARAM_DEFS[i]["paramCode"], PARAM_DEFS[i]["paramName"], rand_time().isoformat()])

    # 缺陷实体
    for i, did in enumerate(defect_type_ids):
        eid = uid()
        entity_map.setdefault("Defect", []).append(eid)
        ws2.append([eid, gvid, "Defect", "qc", "defect_type", did,
                     DEFECT_TYPES[i]["defectCode"], DEFECT_TYPES[i]["defectName"], rand_time().isoformat()])

    # 批次实体
    for i, bid in enumerate(batch_ids[:5]):
        eid = uid()
        entity_map.setdefault("Batch", []).append(eid)
        ws2.append([eid, gvid, "Batch", "prod", "production_batch", bid,
                     BATCH_NO_PREFIXES[i], BATCH_NO_PREFIXES[i], rand_time().isoformat()])

    # KgRelation
    ws3 = wb.create_sheet("kg_relation")
    ws3.append(["relationId", "graphVersionId", "sourceEntityId", "targetEntityId",
                 "relationType", "relationWeight", "confidence", "createdAt"])

    # 工序 -> 参数 (HAS_PARAMETER)
    for step_eid in entity_map.get("ProcessStep", []):
        for param_eid in random.sample(entity_map.get("ProcessParameter", []), min(3, len(entity_map.get("ProcessParameter", [])))):
            ws3.append([uid(), gvid, step_eid, param_eid, "HAS_PARAMETER",
                        rand_value(0.5, 1.0), rand_value(0.8, 1.0), rand_time().isoformat()])

    # 参数 -> 缺陷 (CAUSES_DEFECT)
    for param_eid in entity_map.get("ProcessParameter", []):
        for def_eid in random.sample(entity_map.get("Defect", []), min(2, len(entity_map.get("Defect", [])))):
            ws3.append([uid(), gvid, param_eid, def_eid, "CAUSES_DEFECT",
                        rand_value(0.3, 0.9), rand_value(0.6, 0.95), rand_time().isoformat()])

    # 工序 -> 工序 (NEXT_STEP)
    step_eids = entity_map.get("ProcessStep", [])
    for i in range(len(step_eids) - 1):
        ws3.append([uid(), gvid, step_eids[i], step_eids[i + 1], "NEXT_STEP",
                    1.0, 1.0, rand_time().isoformat()])

    save_workbook(wb, "kg_data.xlsx")


# ============================================================
# 生成评估数据
# ============================================================

def gen_eval_data(batch_ids, step_ids):
    print("生成 eval 模块数据...")
    wb = Workbook()

    # AssessmentTask
    ws = wb.active
    ws.title = "assessment_task"
    ws.append(["taskId", "taskType", "batchId", "stepId", "modelName", "modelVersion",
                 "taskStatus", "createdAt", "finishedAt"])
    task_ids = []
    for i, bid in enumerate(batch_ids[:5]):
        tid = uid()
        task_ids.append(tid)
        ws.append([tid, random.choice(["QUALIFIED", "JUDGMENT", "PREDICTION"]),
                   bid, random.choice(step_ids), "XGBoost", "v1.0", "COMPLETED",
                   rand_time(i).isoformat(), (rand_time(i) + timedelta(minutes=random.randint(5, 30))).isoformat()])

    # AssessmentResult
    ws2 = wb.create_sheet("assessment_result")
    ws2.append(["resultId", "taskId", "assessmentScore", "passProbability", "isPass",
                 "riskLevel", "conclusion", "createdAt"])
    for tid in task_ids:
        score = rand_value(70, 99)
        prob = rand_value(0.6, 0.99, 4)
        is_pass = score >= 85
        ws2.append([uid(), tid, score, prob, is_pass,
                     "低" if score >= 90 else ("中" if score >= 80 else "高"),
                     "合格" if is_pass else "存在风险，建议优化参数",
                     rand_time().isoformat()])

    save_workbook(wb, "eval_data.xlsx")


# ============================================================
# 主函数
# ============================================================

def main():
    print("=" * 50)
    print("电子元器件装配质量评估系统 - 测试数据生成")
    print("=" * 50)

    step_ids, station_ids, equip_ids, product_ids, param_ids = gen_core_data()
    batch_ids, unit_ids, run_ids = gen_prod_data(step_ids, station_ids, equip_ids, product_ids, param_ids)
    defect_type_ids, metric_ids = gen_qc_data(step_ids, unit_ids, run_ids)
    gen_kg_data(step_ids, param_ids, defect_type_ids, batch_ids)
    gen_eval_data(batch_ids, step_ids)

    print("=" * 50)
    print(f"所有文件已生成到 {OUTPUT_DIR}/ 目录")
    print("=" * 50)


if __name__ == "__main__":
    main()
